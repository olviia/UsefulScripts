"""
Google drive backup verification script

Validates every file in the backup, with special handling for .docx and .xlsx:
  - Structural check (python-docx / openpyxl)
  - COM open test via Microsoft Word / Excel
  - Auto-repairs broken Google exports via LibreOffice headless re-save

Install dependencies:
    pip install python-docx openpyxl python-pptx pypdf Pillow pywin32

Requirements:
    - Microsoft Word and Excel (for COM validation)
    - LibreOffice (for auto-repair) — https://www.libreoffice.org/download

Usage:
    python verify_backup.py "C:\YourPath"
    python verify_backup.py "C:\YourPath" --no-word      (skip Office COM checks)
    python verify_backup.py "C:\YourPath" --no-repair    (skip auto-repair)
"""

import os
import sys
import csv
import time
import shutil
import subprocess
import argparse
from datetime import datetime
from pathlib import Path

# ── Optional imports ─────────────────────────────────────────────────────────
VALIDATORS = {}
WORD_APP = None
EXCEL_APP = None
USE_WORD_COM = False
USE_EXCEL_COM = False
LIBREOFFICE_PATH = None

# Detect LibreOffice
for lo_path in [
    r"C:\Program Files\LibreOffice\program\soffice.exe",
    r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
    "/usr/bin/libreoffice",
    "/usr/local/bin/libreoffice",
]:
    if os.path.exists(lo_path):
        LIBREOFFICE_PATH = lo_path
        break

if LIBREOFFICE_PATH:
    print(f"LibreOffice found: {LIBREOFFICE_PATH}")
else:
    print("WARNING: LibreOffice not found — auto-repair will be unavailable")
    print("         Install from https://www.libreoffice.org/download")

try:
    from docx import Document as DocxDocument
    def validate_docx_basic(path):
        doc = DocxDocument(path)
        _ = len(doc.paragraphs)
    VALIDATORS['.docx_basic'] = validate_docx_basic
except ImportError:
    print("WARNING: python-docx not installed, .docx basic check unavailable")

try:
    import win32com.client
    import pythoncom

    def init_word():
        global WORD_APP
        pythoncom.CoInitialize()
        WORD_APP = win32com.client.DispatchEx('Word.Application')
        WORD_APP.Visible = False
        WORD_APP.DisplayAlerts = 0

    def init_excel():
        global EXCEL_APP
        if not EXCEL_APP:
            EXCEL_APP = win32com.client.DispatchEx('Excel.Application')
            EXCEL_APP.Visible = False
            EXCEL_APP.DisplayAlerts = False

    def shutdown_word():
        global WORD_APP
        if WORD_APP:
            try:
                WORD_APP.Quit()
            except:
                pass
            WORD_APP = None

    def shutdown_excel():
        global EXCEL_APP
        if EXCEL_APP:
            try:
                EXCEL_APP.Quit()
            except:
                pass
            EXCEL_APP = None

    def shutdown_com():
        shutdown_word()
        shutdown_excel()
        pythoncom.CoUninitialize()

    def validate_docx_word(path):
        global WORD_APP
        abs_path = os.path.abspath(path)
        doc = None
        try:
            doc = WORD_APP.Documents.Open(
                abs_path,
                ConfirmConversions=False,
                ReadOnly=True,
                AddToRecentFiles=False,
                OpenAndRepair=False
            )
            page_count = doc.ComputeStatistics(2)
        finally:
            if doc:
                try:
                    doc.Close(SaveChanges=0)
                except:
                    pass

    def validate_xlsx_excel(path):
        global EXCEL_APP
        abs_path = os.path.abspath(path)
        wb = None
        try:
            wb = EXCEL_APP.Workbooks.Open(
                abs_path,
                ReadOnly=True,
                UpdateLinks=0,
                CorruptLoad=1  # xlNormalLoad
            )
            _ = wb.Sheets.Count
        finally:
            if wb:
                try:
                    wb.Close(SaveChanges=False)
                except:
                    pass

    USE_WORD_COM = True
    USE_EXCEL_COM = True
    print("Word and Excel COM automation available")
except ImportError:
    print("WARNING: pywin32 not installed — install with: pip install pywin32")
    print("         Falling back to python-docx only (won't catch Google export issues)")

try:
    from openpyxl import load_workbook
    def validate_xlsx_basic(path):
        wb = load_workbook(path, read_only=True)
        _ = wb.sheetnames
        wb.close()
    VALIDATORS['.xlsx_basic'] = validate_xlsx_basic
except ImportError:
    print("WARNING: openpyxl not installed")

try:
    from pptx import Presentation
    def validate_pptx(path):
        prs = Presentation(path)
        _ = len(prs.slides)
    VALIDATORS['.pptx'] = validate_pptx
except ImportError:
    print("WARNING: python-pptx not installed")

try:
    from pypdf import PdfReader
    def validate_pdf(path):
        reader = PdfReader(path)
        _ = len(reader.pages)
    VALIDATORS['.pdf'] = validate_pdf
except ImportError:
    print("WARNING: pypdf not installed")

try:
    from PIL import Image
    IMAGE_EXTS = {'.png', '.jpg', '.jpeg', '.gif', '.bmp', '.tiff', '.tif', '.webp'}
    def validate_image(path):
        img = Image.open(path)
        img.verify()
    for ext in IMAGE_EXTS:
        VALIDATORS[ext] = validate_image
except ImportError:
    print("WARNING: Pillow not installed")


# ── LibreOffice repair ───────────────────────────────────────────────────────

def repair_via_libreoffice(filepath, target_format):
    """
    Re-save a file through LibreOffice to fix Google export issues.
    Creates a .bak backup before modifying.
    target_format: 'docx' or 'xlsx'
    Returns (success, detail).
    """
    if not LIBREOFFICE_PATH:
        return False, "LibreOffice not available"

    abs_path = os.path.abspath(filepath)
    parent_dir = os.path.dirname(abs_path)
    filename = os.path.basename(abs_path)
    backup_path = abs_path + '.bak'

    # Create backup
    try:
        shutil.copy2(abs_path, backup_path)
    except Exception as e:
        return False, f"Failed to create backup: {e}"

    # LibreOffice converts to a temp directory to avoid overwrite issues
    tmp_dir = os.path.join(parent_dir, '__lo_repair_tmp__')
    os.makedirs(tmp_dir, exist_ok=True)

    try:
        result = subprocess.run(
            [LIBREOFFICE_PATH, '--headless', '--convert-to', target_format,
             '--outdir', tmp_dir, abs_path],
            capture_output=True, text=True, timeout=120
        )

        if result.returncode != 0:
            return False, f"LibreOffice failed: {result.stderr}"

        # Find the output file
        repaired_path = os.path.join(tmp_dir, filename)
        if not os.path.exists(repaired_path):
            for f in os.listdir(tmp_dir):
                if f.endswith(f'.{target_format}'):
                    repaired_path = os.path.join(tmp_dir, f)
                    break

        if not os.path.exists(repaired_path):
            return False, "LibreOffice produced no output"

        # Replace original with repaired
        shutil.move(repaired_path, abs_path)
        return True, "Repaired via LibreOffice re-save"

    except subprocess.TimeoutExpired:
        return False, "LibreOffice timed out (120s)"
    except Exception as e:
        # Restore backup on failure
        if os.path.exists(backup_path):
            shutil.copy2(backup_path, abs_path)
        return False, f"Repair failed: {type(e).__name__}: {e}"
    finally:
        # Clean up temp dir
        try:
            shutil.rmtree(tmp_dir, ignore_errors=True)
        except:
            pass


# ── Validation ───────────────────────────────────────────────────────────────

def validate_file(filepath, use_word=True, auto_repair=True):
    """
    Validate a single file. Returns (status, detail).

    For .docx:
      1. python-docx structural check
      2. Word COM open test (if enabled)
      3. If Word fails → LibreOffice re-save → re-test
    """
    path = Path(filepath)

    if not path.exists():
        return "MISSING", "File does not exist"

    size = path.stat().st_size
    if size == 0:
        return "EMPTY", "File is 0 bytes"

    ext = path.suffix.lower()

    # ── Special .docx handling ───────────────────────────────────────────
    if ext == '.docx':
        basic_ok = True
        basic_detail = ""
        word_ok = True
        word_detail = ""
        repair_detail = ""

        # python-docx check
        basic_validator = VALIDATORS.get('.docx_basic')
        if basic_validator:
            try:
                basic_validator(filepath)
                basic_detail = "python-docx OK"
            except Exception as e:
                basic_ok = False
                basic_detail = f"python-docx FAILED: {type(e).__name__}: {e}"

        # Word COM check
        if use_word and USE_WORD_COM and WORD_APP:
            try:
                validate_docx_word(filepath)
                word_detail = "Word COM OK"
            except Exception as e:
                word_ok = False
                word_detail = f"Word COM FAILED: {type(e).__name__}: {e}"

                # ── Auto-repair via LibreOffice ──────────────────────────
                if auto_repair and LIBREOFFICE_PATH:
                    success, rep_detail = repair_via_libreoffice(filepath, 'docx')
                    if success:
                        repair_detail = rep_detail
                        # Re-test with Word
                        try:
                            validate_docx_word(filepath)
                            word_ok = True
                            word_detail = f"Word COM OK (after LibreOffice repair)"
                            # Remove .bak since repair succeeded
                            bak_path = filepath + '.bak'
                            if os.path.exists(bak_path):
                                os.remove(bak_path)
                        except Exception as e2:
                            word_detail = f"Word COM STILL FAILS after repair: {type(e2).__name__}: {e2}"
                            # Restore original from .bak since repair didn't help
                            bak_path = filepath + '.bak'
                            if os.path.exists(bak_path):
                                shutil.copy2(bak_path, filepath)
                                os.remove(bak_path)
                    else:
                        repair_detail = f"Repair attempted but failed: {rep_detail}"

        detail = f"{size:,} bytes | {basic_detail} | {word_detail}"
        if repair_detail:
            detail += f" | {repair_detail}"

        if not basic_ok and not word_ok:
            return "CORRUPT", detail
        elif basic_ok and not word_ok:
            return "WORD_FAIL", detail
        elif not basic_ok and word_ok:
            return "PARSE_WARN", detail
        else:
            if "after LibreOffice repair" in word_detail:
                return "REPAIRED", detail
            return "OK", detail

    # ── Special .xlsx handling ───────────────────────────────────────────
    if ext == '.xlsx':
        basic_ok = True
        basic_detail = ""
        excel_ok = True
        excel_detail = ""
        repair_detail = ""

        # openpyxl check
        basic_validator = VALIDATORS.get('.xlsx_basic')
        if basic_validator:
            try:
                basic_validator(filepath)
                basic_detail = "openpyxl OK"
            except Exception as e:
                basic_ok = False
                basic_detail = f"openpyxl FAILED: {type(e).__name__}: {e}"

        # Excel COM check
        if use_word and USE_EXCEL_COM and EXCEL_APP:
            try:
                validate_xlsx_excel(filepath)
                excel_detail = "Excel COM OK"
            except Exception as e:
                excel_ok = False
                excel_detail = f"Excel COM FAILED: {type(e).__name__}: {e}"

                # ── Auto-repair via LibreOffice ──────────────────────────
                if auto_repair and LIBREOFFICE_PATH:
                    success, rep_detail = repair_via_libreoffice(filepath, 'xlsx')
                    if success:
                        repair_detail = rep_detail
                        try:
                            validate_xlsx_excel(filepath)
                            excel_ok = True
                            excel_detail = f"Excel COM OK (after LibreOffice repair)"
                            bak_path = filepath + '.bak'
                            if os.path.exists(bak_path):
                                os.remove(bak_path)
                        except Exception as e2:
                            excel_detail = f"Excel COM STILL FAILS after repair: {type(e2).__name__}: {e2}"
                            bak_path = filepath + '.bak'
                            if os.path.exists(bak_path):
                                shutil.copy2(bak_path, filepath)
                                os.remove(bak_path)
                    else:
                        repair_detail = f"Repair attempted but failed: {rep_detail}"

        detail = f"{size:,} bytes | {basic_detail} | {excel_detail}"
        if repair_detail:
            detail += f" | {repair_detail}"

        if not basic_ok and not excel_ok:
            return "CORRUPT", detail
        elif basic_ok and not excel_ok:
            return "EXCEL_FAIL", detail
        elif not basic_ok and excel_ok:
            return "PARSE_WARN", detail
        else:
            if "after LibreOffice repair" in excel_detail:
                return "REPAIRED", detail
            return "OK", detail

    # ── Standard validators ──────────────────────────────────────────────
    validator = VALIDATORS.get(ext)
    if validator:
        try:
            validator(filepath)
            return "OK", f"{size:,} bytes, content validated"
        except Exception as e:
            return "CORRUPT", f"{size:,} bytes, validation failed: {type(e).__name__}: {e}"
    else:
        try:
            with open(filepath, 'rb') as f:
                f.read(1024)
            return "OK", f"{size:,} bytes, readable (no format validator)"
        except Exception as e:
            return "ERROR", f"Cannot read file: {e}"


# ── Main ─────────────────────────────────────────────────────────────────────

def verify_backup(backup_dir, use_word=True, auto_repair=True):
    backup_path = Path(backup_dir)
    if not backup_path.exists():
        print(f"ERROR: Directory not found: {backup_dir}")
        sys.exit(1)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file = backup_path.parent / f"verification_log_{timestamp}.csv"

    all_files = []
    for root, dirs, files in os.walk(backup_path):
        for fname in files:
            if not fname.endswith('.bak'):
                all_files.append(os.path.join(root, fname))

    total = len(all_files)
    docx_count = sum(1 for f in all_files if f.lower().endswith('.docx'))
    xlsx_count = sum(1 for f in all_files if f.lower().endswith('.xlsx'))
    office_count = docx_count + xlsx_count

    print(f"\nFound {total:,} files to verify in: {backup_dir}")
    print(f"  .docx files:          {docx_count:,}")
    print(f"  .xlsx files:          {xlsx_count:,}")
    print(f"  Office COM validation: {'ENABLED' if use_word and (USE_WORD_COM or USE_EXCEL_COM) else 'DISABLED'}")
    print(f"  Auto-repair:          {'ENABLED (LibreOffice)' if auto_repair and LIBREOFFICE_PATH else 'DISABLED'}")
    print(f"  Log:                  {log_file}\n")

    if use_word and USE_WORD_COM and docx_count > 0:
        print("Starting Microsoft Word...")
        try:
            init_word()
            print("Word started.")
        except Exception as e:
            print(f"Failed to start Word: {e}")

    if use_word and USE_EXCEL_COM and xlsx_count > 0:
        print("Starting Microsoft Excel...")
        try:
            init_excel()
            print("Excel started.")
        except Exception as e:
            print(f"Failed to start Excel: {e}")

    if office_count > 0:
        print()

    stats = {"OK": 0, "EMPTY": 0, "CORRUPT": 0, "ERROR": 0, "MISSING": 0,
             "WORD_FAIL": 0, "EXCEL_FAIL": 0, "PARSE_WARN": 0, "REPAIRED": 0}
    start_time = time.time()

    try:
        with open(log_file, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(["Timestamp", "Status", "Relative Path", "Extension", "Detail"])

            for i, filepath in enumerate(all_files, 1):
                rel_path = os.path.relpath(filepath, backup_path)
                ext = Path(filepath).suffix.lower()

                status, detail = validate_file(filepath, use_word=use_word, auto_repair=auto_repair)
                stats[status] = stats.get(status, 0) + 1

                now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                writer.writerow([now, status, rel_path, ext, detail])

                if i % 50 == 0 or i == total or status not in ("OK",):
                    elapsed = time.time() - start_time
                    rate = i / elapsed if elapsed > 0 else 0
                    remaining = (total - i) / rate if rate > 0 else 0

                    if status == "OK":
                        print(f"  [{i:,}/{total:,}] OK        | ETA: {remaining:.0f}s")
                    else:
                        print(f"  [{i:,}/{total:,}] {status:9s} | {rel_path[:70]}")

    finally:
        if WORD_APP or EXCEL_APP:
            print("\nClosing Office applications...")
            shutdown_com()

    elapsed_total = time.time() - start_time

    print(f"\n{'='*60}")
    print(f"VERIFICATION COMPLETE")
    print(f"{'='*60}")
    print(f"Total files:    {total:,}")
    print(f"OK:             {stats['OK']:,}")
    print(f"Repaired:       {stats['REPAIRED']:,}  (auto-fixed via LibreOffice)")
    print(f"Empty:          {stats['EMPTY']:,}")
    print(f"Corrupt:        {stats['CORRUPT']:,}")
    print(f"Word failures:  {stats['WORD_FAIL']:,}  (could not be auto-repaired)")
    print(f"Excel failures: {stats['EXCEL_FAIL']:,}  (could not be auto-repaired)")
    print(f"Parse warnings: {stats['PARSE_WARN']:,}")
    print(f"Errors:         {stats['ERROR']:,}")
    print(f"Time:           {elapsed_total:.1f}s")
    print(f"Log saved:      {log_file}")
    print(f"{'='*60}")

    problems = stats['CORRUPT'] + stats['EMPTY'] + stats['ERROR'] + stats['WORD_FAIL'] + stats['EXCEL_FAIL']
    if problems > 0:
        print(f"\n⚠ {problems} files need attention. Check the CSV log.")
        if stats['WORD_FAIL'] > 0:
            print(f"  → {stats['WORD_FAIL']} .docx files could not be repaired automatically.")
        if stats['EXCEL_FAIL'] > 0:
            print(f"  → {stats['EXCEL_FAIL']} .xlsx files could not be repaired automatically.")
        return 1
    else:
        if stats['REPAIRED'] > 0:
            print(f"\n✓ All files OK ({stats['REPAIRED']} were auto-repaired via LibreOffice).")
        else:
            print(f"\n✓ All files verified successfully.")
        return 0


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Verify backup file integrity")
    parser.add_argument("backup_dir", help="Path to backup directory")
    parser.add_argument("--no-word", action="store_true", help="Skip Word COM validation")
    parser.add_argument("--no-repair", action="store_true", help="Skip auto-repair")
    args = parser.parse_args()

    sys.exit(verify_backup(args.backup_dir,
                           use_word=not args.no_word,
                           auto_repair=not args.no_repair))
